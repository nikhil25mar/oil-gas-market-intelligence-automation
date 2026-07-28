import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    summary = pd.read_csv("daily_summary.csv")
    if summary.empty:
        raise pd.errors.EmptyDataError
except (FileNotFoundError, pd.errors.EmptyDataError):
    print("No data available yet for today's report (not enough price history). Skipping report generation.")
    exit()

try:
    with open("market_commentary.txt", "r", encoding="utf-8") as f:
        commentary_text = f.read()
except FileNotFoundError:
    commentary_text = ""

wb = Workbook()
ws = wb.active
ws.title = "Daily Report"

# Title row
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = f"Oil & Gas Market Intelligence Report — {datetime.now().strftime('%Y-%m-%d')}"
title_cell.font = Font(size=14, bold=True, color="FFFFFF")
title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center")

# Write header + data starting row 3
start_row = 3
for r_idx, row in enumerate(dataframe_to_rows(summary, index=False, header=True), start=start_row):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == start_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

# Highlight alert rows in red text
alert_col_index = summary.columns.get_loc("Alert") + 1
for r_idx in range(start_row + 1, start_row + 1 + len(summary)):
    alert_value = ws.cell(row=r_idx, column=alert_col_index).value
    if alert_value and "BIG MOVE" in str(alert_value):
        for c_idx in range(1, len(summary.columns) + 1):
            ws.cell(row=r_idx, column=c_idx).font = Font(color="C00000", bold=True)

# Auto-width columns (approximate) — based on header + data rows only, skips merged title row
for col_idx in range(1, len(summary.columns) + 1):
    max_length = 0
    for r_idx in range(start_row, start_row + 1 + len(summary)):
        cell_value = ws.cell(row=r_idx, column=col_idx).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = max_length + 4

# Add commentary section below the data table (BEFORE saving)
commentary_start_row = start_row + len(summary) + 3

ws.merge_cells(f"A{commentary_start_row}:G{commentary_start_row}")
header_cell = ws[f"A{commentary_start_row}"]
header_cell.value = "Market Commentary"
header_cell.font = Font(size=12, bold=True, color="FFFFFF")
header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

commentary_row = commentary_start_row + 1
for line in commentary_text.split("\n"):
    ws.merge_cells(f"A{commentary_row}:G{commentary_row}")
    cell = ws[f"A{commentary_row}"]
    cell.value = line
    cell.alignment = Alignment(wrap_text=True)
    commentary_row += 1

# Save AFTER everything (data + commentary) has been added
output_file = f"OilGas_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(output_file)
print(f"Report saved as {output_file}")